"""
pivot_engine.py
===============

Turn a raw multi-channel marketing-campaign CSV into a formatted, multi-sheet
Excel workbook of the summaries a marketing team actually asks for.

This file is organized in three readable layers:

  1. METRIC FUNCTIONS  -- pure pandas: take a DataFrame, return a summary.
                          No file reading, no Excel. Easy to read and to test.
  2. WORKBOOK WRITER   -- takes those summaries and formats them with openpyxl.
  3. COMMAND LINE      -- glues 1 and 2 together: read CSV -> compute -> write.

Usage:
    python pivot_engine.py demo.csv
    python pivot_engine.py demo.csv -o my_report.xlsx
"""

import argparse
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Metric definitions -- ONE place, reused in the workbook so a reviewer always
# sees exactly how each number was calculated.
# --------------------------------------------------------------------------- #
METRIC_DEFINITIONS = {
    "ROAS": "Return on ad spend = revenue / spend",
    "CTR": "Click-through rate = clicks / impressions",
    "Conversion rate": "conversions / clicks",
    "CPA": "Cost per acquisition = spend / conversions",
    "Revenue per conversion": "revenue / conversions",
}


# --------------------------------------------------------------------------- #
# Layer 1: metric functions (pure -- no I/O)
# --------------------------------------------------------------------------- #
def _safe_div(numerator, denominator):
    """Divide, but return NaN instead of infinity when dividing by zero.

    Works for single numbers and for whole pandas columns (Series). This keeps
    an empty channel/segment from blowing up the report with 'inf'.
    """
    if isinstance(numerator, pd.Series) or isinstance(denominator, pd.Series):
        with np.errstate(divide="ignore", invalid="ignore"):
            result = numerator / denominator
        return result.replace([np.inf, -np.inf], np.nan)
    if denominator == 0:
        return float("nan")
    return numerator / denominator


def channel_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Spend, revenue, ROAS, conversions and CPA grouped by channel.

    Sorted best-to-worst by ROAS so the top of the table is the winner.
    """
    g = df.groupby("channel", as_index=False).agg(
        spend=("spend", "sum"),
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        conversions=("conversions", "sum"),
        revenue=("revenue", "sum"),
    )
    g["roas"] = _safe_div(g["revenue"], g["spend"])
    g["cpa"] = _safe_div(g["spend"], g["conversions"])
    g["conversion_rate"] = _safe_div(g["conversions"], g["clicks"])
    return g.sort_values("roas", ascending=False, ignore_index=True)


def segment_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Clicks, conversions, conversion rate and revenue grouped by segment.

    Sorted best-to-worst by conversion rate.
    """
    g = df.groupby("segment", as_index=False).agg(
        clicks=("clicks", "sum"),
        conversions=("conversions", "sum"),
        revenue=("revenue", "sum"),
        spend=("spend", "sum"),
    )
    g["conversion_rate"] = _safe_div(g["conversions"], g["clicks"])
    g["revenue_per_conversion"] = _safe_div(g["revenue"], g["conversions"])
    g["roas"] = _safe_div(g["revenue"], g["spend"])
    return g.sort_values("conversion_rate", ascending=False, ignore_index=True)


def monthly_trend(df: pd.DataFrame) -> pd.DataFrame:
    """Month-by-month spend, revenue, ROAS and the month-over-month change.

    `mom_revenue_change` is the fractional change in revenue vs the prior month
    (e.g. 0.10 = up 10%). The first month is NaN because it has no prior month.
    """
    d = df.copy()
    d["month"] = pd.to_datetime(d["date"]).dt.strftime("%Y-%m")
    g = d.groupby("month", as_index=False).agg(
        spend=("spend", "sum"),
        revenue=("revenue", "sum"),
        clicks=("clicks", "sum"),
        conversions=("conversions", "sum"),
    )
    g = g.sort_values("month", ignore_index=True)
    g["roas"] = _safe_div(g["revenue"], g["spend"])
    g["mom_revenue_change"] = g["revenue"].pct_change()
    return g


def top_bottom_performers(df: pd.DataFrame, n: int = 5):
    """Rank every channel x segment combination by ROAS.

    Returns a (top_n, bottom_n) tuple of DataFrames -- the best and worst
    performing combinations, each sorted from highest ROAS to lowest within
    its group.
    """
    g = df.groupby(["channel", "segment"], as_index=False).agg(
        spend=("spend", "sum"),
        clicks=("clicks", "sum"),
        conversions=("conversions", "sum"),
        revenue=("revenue", "sum"),
    )
    g["roas"] = _safe_div(g["revenue"], g["spend"])
    g = g.sort_values(["roas", "revenue"], ascending=False, ignore_index=True)
    top = g.head(n).reset_index(drop=True)
    bottom = (
        g.tail(n)
        .sort_values(["roas", "revenue"], ascending=False, ignore_index=True)
    )
    return top, bottom


def compute_kpis(df: pd.DataFrame) -> dict:
    """One-page headline numbers for the whole dataset."""
    total_spend = df["spend"].sum()
    total_revenue = df["revenue"].sum()
    total_impressions = df["impressions"].sum()
    total_clicks = df["clicks"].sum()
    total_conversions = df["conversions"].sum()

    channels = channel_summary(df)
    segments = segment_summary(df)

    return {
        "total_spend": total_spend,
        "total_revenue": total_revenue,
        "total_impressions": total_impressions,
        "total_clicks": total_clicks,
        "total_conversions": total_conversions,
        "overall_roas": _safe_div(total_revenue, total_spend),
        "overall_ctr": _safe_div(total_clicks, total_impressions),
        "conversion_rate": _safe_div(total_conversions, total_clicks),
        "cpa": _safe_div(total_spend, total_conversions),
        "revenue_per_conversion": _safe_div(total_revenue, total_conversions),
        "best_channel_by_roas": channels.loc[channels["roas"].idxmax(), "channel"],
        "worst_channel_by_roas": channels.loc[channels["roas"].idxmin(), "channel"],
        "best_segment_by_conversion": segments.loc[
            segments["conversion_rate"].idxmax(), "segment"
        ],
    }


# --------------------------------------------------------------------------- #
# Layer 2: workbook writer (openpyxl formatting)
# --------------------------------------------------------------------------- #

# Excel number-format codes.
CURRENCY = '"$"#,##0'
RATIO = "0.00"
PERCENT = "0.00%"
INTEGER = "#,##0"

# Sheet names -- the workbook is built in this exact order.
SHEET_KPI = "KPI Summary"
SHEET_CHANNEL = "Revenue & ROAS by Channel"
SHEET_SEGMENT = "Conversion Rate by Segment"
SHEET_TREND = "Month-over-Month Trend"
SHEET_PERFORMERS = "Top-Bottom 5 Performers"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_NOTE_FONT = Font(italic=True, color="808080")
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_CENTER = Alignment(horizontal="center")


def _style_header_cell(cell):
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER


def _to_native(value):
    """Convert numpy / NaN values into things openpyxl writes cleanly."""
    if pd.isna(value):
        return None
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    return value


def _write_table(ws, df, col_specs, start_row=1):
    """Write a DataFrame as a formatted table on worksheet `ws`.

    col_specs is a list of (dataframe_column, header_label, number_format_or_None).
    Returns (first_data_row, last_data_row) so the caller can add conditional
    formatting to specific columns afterwards.
    """
    header_row = start_row
    for col_idx, (_, header, _) in enumerate(col_specs, start=1):
        _style_header_cell(ws.cell(row=header_row, column=col_idx, value=header))

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + i
        for col_idx, (col, _, fmt) in enumerate(col_specs, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=_to_native(row[col]))
            if fmt and cell.value is not None:
                cell.number_format = fmt

    return header_row + 1, header_row + len(df)


def _add_roas_rule(ws, col_index, first_row, last_row):
    """Shade a ROAS column: green when >= 1 (profitable), red when < 1."""
    if last_row < first_row:
        return
    col = get_column_letter(col_index)
    cells = f"{col}{first_row}:{col}{last_row}"
    ws.conditional_formatting.add(
        cells, CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=_GREEN_FILL)
    )
    ws.conditional_formatting.add(
        cells, CellIsRule(operator="lessThan", formula=["1"], fill=_RED_FILL)
    )


def _autosize(ws, min_width=10, max_width=44):
    for column_cells in ws.columns:
        length = max(
            (len(str(c.value)) for c in column_cells if c.value is not None),
            default=0,
        )
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def _build_kpi_sheet(ws, df, source_name):
    kpis = compute_kpis(df)
    ws["A1"] = "Campaign KPI Summary"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"SYNTHETIC demo data  -  source: {source_name}  ({len(df):,} rows)"
    ws["A2"].font = _NOTE_FONT

    table_header = 4
    _style_header_cell(ws.cell(row=table_header, column=1, value="Metric"))
    _style_header_cell(ws.cell(row=table_header, column=2, value="Value"))

    kpi_rows = [
        ("Total Spend", kpis["total_spend"], CURRENCY),
        ("Total Revenue", kpis["total_revenue"], CURRENCY),
        ("Overall ROAS", kpis["overall_roas"], RATIO),
        ("Total Conversions", kpis["total_conversions"], INTEGER),
        ("Conversion Rate", kpis["conversion_rate"], PERCENT),
        ("Total Clicks", kpis["total_clicks"], INTEGER),
        ("Total Impressions", kpis["total_impressions"], INTEGER),
        ("Overall CTR", kpis["overall_ctr"], PERCENT),
        ("CPA (cost per conversion)", kpis["cpa"], CURRENCY),
        ("Revenue per Conversion", kpis["revenue_per_conversion"], CURRENCY),
        ("Best Channel (by ROAS)", kpis["best_channel_by_roas"], None),
        ("Worst Channel (by ROAS)", kpis["worst_channel_by_roas"], None),
        ("Best Segment (by Conversion Rate)", kpis["best_segment_by_conversion"], None),
    ]
    roas_row = None
    for offset, (label, value, fmt) in enumerate(kpi_rows, start=1):
        r = table_header + offset
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=_to_native(value))
        if fmt and cell.value is not None:
            cell.number_format = fmt
        if label == "Overall ROAS":
            roas_row = r
    _add_roas_rule(ws, 2, roas_row, roas_row)

    defs_start = table_header + len(kpi_rows) + 2
    ws.cell(row=defs_start, column=1, value="How these metrics are defined").font = _BOLD
    for k, (name, definition) in enumerate(METRIC_DEFINITIONS.items(), start=defs_start + 1):
        ws.cell(row=k, column=1, value=name)
        ws.cell(row=k, column=2, value=definition)

    _autosize(ws)
    ws.freeze_panes = "A5"


def _build_channel_sheet(ws, df):
    specs = [
        ("channel", "Channel", None),
        ("spend", "Spend", CURRENCY),
        ("revenue", "Revenue", CURRENCY),
        ("roas", "ROAS", RATIO),
        ("conversions", "Conversions", INTEGER),
        ("cpa", "CPA", CURRENCY),
        ("conversion_rate", "Conversion Rate", PERCENT),
    ]
    first, last = _write_table(ws, channel_summary(df), specs)
    _add_roas_rule(ws, 4, first, last)          # ROAS is the 4th column
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_segment_sheet(ws, df):
    specs = [
        ("segment", "Segment", None),
        ("clicks", "Clicks", INTEGER),
        ("conversions", "Conversions", INTEGER),
        ("conversion_rate", "Conversion Rate", PERCENT),
        ("revenue", "Revenue", CURRENCY),
        ("revenue_per_conversion", "Revenue / Conversion", CURRENCY),
        ("roas", "ROAS", RATIO),
    ]
    first, last = _write_table(ws, segment_summary(df), specs)
    _add_roas_rule(ws, 7, first, last)          # ROAS is the 7th column
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_trend_sheet(ws, df):
    specs = [
        ("month", "Month", None),
        ("spend", "Spend", CURRENCY),
        ("revenue", "Revenue", CURRENCY),
        ("roas", "ROAS", RATIO),
        ("conversions", "Conversions", INTEGER),
        ("mom_revenue_change", "MoM Revenue Change", PERCENT),
    ]
    first, last = _write_table(ws, monthly_trend(df), specs)
    _add_roas_rule(ws, 4, first, last)
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_performers_sheet(ws, df, n=5):
    top, bottom = top_bottom_performers(df, n=n)
    specs = [
        ("channel", "Channel", None),
        ("segment", "Segment", None),
        ("spend", "Spend", CURRENCY),
        ("revenue", "Revenue", CURRENCY),
        ("roas", "ROAS", RATIO),
        ("conversions", "Conversions", INTEGER),
    ]
    ws.cell(row=1, column=1, value=f"Top {n} performers (by ROAS)").font = _BOLD
    first, last = _write_table(ws, top, specs, start_row=2)
    _add_roas_rule(ws, 5, first, last)          # ROAS is the 5th column

    gap = last + 2
    ws.cell(row=gap, column=1, value=f"Bottom {n} performers (by ROAS)").font = _BOLD
    first, last = _write_table(ws, bottom, specs, start_row=gap + 1)
    _add_roas_rule(ws, 5, first, last)
    _autosize(ws)


def write_workbook(df: pd.DataFrame, output_path, source_name: str = "campaign data") -> list:
    """Build and save the formatted, multi-sheet workbook. Returns the sheet names."""
    wb = Workbook()
    _build_kpi_sheet(wb.active, df, source_name)
    wb.active.title = SHEET_KPI
    _build_channel_sheet(wb.create_sheet(SHEET_CHANNEL), df)
    _build_segment_sheet(wb.create_sheet(SHEET_SEGMENT), df)
    _build_trend_sheet(wb.create_sheet(SHEET_TREND), df)
    _build_performers_sheet(wb.create_sheet(SHEET_PERFORMERS), df)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return [ws.title for ws in wb.worksheets]


# --------------------------------------------------------------------------- #
# Layer 3: command line
# --------------------------------------------------------------------------- #
REQUIRED_COLUMNS = [
    "date", "channel", "segment",
    "spend", "impressions", "clicks", "conversions", "revenue",
]

DEFAULT_OUTPUT = Path(__file__).resolve().parent / "output" / "campaign_summary.xlsx"


def load_campaign_csv(path) -> pd.DataFrame:
    """Read the campaign CSV and fail loudly if a required column is missing."""
    df = pd.read_csv(path)
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise SystemExit(
            f"Input CSV is missing required column(s): {', '.join(missing)}\n"
            f"Expected columns: {', '.join(REQUIRED_COLUMNS)}"
        )
    return df


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Turn a campaign CSV into a formatted Excel summary workbook.",
    )
    parser.add_argument("csv_path", type=Path, help="path to the input campaign CSV")
    parser.add_argument(
        "-o", "--output", type=Path, default=DEFAULT_OUTPUT,
        help=f"output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    start = time.perf_counter()
    df = load_campaign_csv(args.csv_path)
    sheets = write_workbook(df, args.output, source_name=Path(args.csv_path).name)
    elapsed = time.perf_counter() - start
    print(f"Wrote {len(sheets)} sheets to {args.output} in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
