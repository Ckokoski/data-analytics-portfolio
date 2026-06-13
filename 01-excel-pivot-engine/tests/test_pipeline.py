"""End-to-end smoke test: generate data, run the CLI, inspect the workbook.

This is the acceptance check from the brief: the whole pipeline must run in
under 30 seconds and produce the five expected sheets.
"""
import subprocess
import sys
import time
import pathlib

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]

EXPECTED_SHEETS = [
    "KPI Summary",
    "Revenue & ROAS by Channel",
    "Conversion Rate by Segment",
    "Month-over-Month Trend",
    "Top-Bottom 5 Performers",
]


def test_end_to_end_under_30s(tmp_path):
    csv = tmp_path / "demo.csv"
    xlsx = tmp_path / "out.xlsx"

    subprocess.run(
        [sys.executable, str(ROOT / "generate_demo_data.py"), "-o", str(csv)],
        check=True,
    )

    start = time.time()
    subprocess.run(
        [sys.executable, str(ROOT / "pivot_engine.py"), str(csv), "-o", str(xlsx)],
        check=True,
    )
    elapsed = time.time() - start
    assert elapsed < 30, f"pipeline took {elapsed:.1f}s (limit 30s)"

    wb = load_workbook(xlsx)
    assert wb.sheetnames == EXPECTED_SHEETS
    # Channel sheet: 1 header row + 4 channels.
    assert wb["Revenue & ROAS by Channel"].max_row == 5
    # KPI sheet carries its title.
    assert wb["KPI Summary"]["A1"].value == "Campaign KPI Summary"
